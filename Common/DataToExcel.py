# coding:utf-8

import xlrd
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter

from openpyxl.reader.excel import load_workbook


def set_style(name, height, bold=False):
    style = xlwt.XFStyle()  # 初始化样式
    font = xlwt.Font()  # 为样式创建字体
    font.name = name  # 'Times New Roman'
    font.bold = bold
    font.color_index = 4
    font.height = height
    borders = xlwt.Borders()
    borders.left = 6
    borders.right = 6
    borders.top = 6
    borders.bottom = 6
    style.font = font
    style.borders = borders
    return style

def write_to_excel_with_openpyxl(df, columnHeaders, filepath="save.xlsx",pageSize=10000):
    # 设置文件输出路径与名称
    # 新建一个workbook
    wb = Workbook()
    from zipfile import ZipFile, ZIP_DEFLATED
    archive = ZipFile(filepath, 'w', ZIP_DEFLATED)
    # 新建一个excelWriter
    ew = ExcelWriter(workbook=wb, archive=archive)

    # 第一个sheet是ws
    ws = wb.worksheets[0]
    # 设置ws的名称
    ws.title = "range names"
    length = len(columnHeaders)
    # 写第一行，标题行
    for i in range(1, length + 1):
        value = columnHeaders[i - 1]
        ws.cell(1, i, value)


    # 写第二行及其以后的那些行
    i = 2
    for index, row in df.iterrows():
        for j in range(1, length + 1):
            # ws.cell(i, j, row[j - 1])
            ws.cell(row=i,column=j).value=str(row[j-1]).encode("utf-8",errors="ignore")
        if i % pageSize == 0:
            ew.save()
            archive = ZipFile(filepath, 'w', ZIP_DEFLATED)
            ew = ExcelWriter(workbook=wb, archive=archive)
        i += 1
    if i%pageSize!=0:
        # 写文件
        ew.save()
    print("导出成功:"+filepath)


def read_excel_with_openpyxl(excel_name="test.xlsx"):
    # 读取excel2007文件
    wb = load_workbook(filename=excel_name)
    # 显示有多少张表
    # print("Worksheet range(s):", wb.get_named_ranges())
    # print("Worksheet name(s):", wb.get_sheet_names())
    # 取第一张表
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # 显示表名，表行数，表列数
    # 获取读入的excel表格的有多少行，有多少列
    row_num = ws.max_row
    col_num = ws.max_column
    print("Title:", ws.title, "row_num: ", row_num, " col_num: ", col_num)
    print("--------------------------------")
    # 建立存储数据的字典
    data_dic = {}
    head=[]
    sign = 1
    # 把数据存到字典中
    for row in ws.rows:
        temp_list = []
        # print("row:", row)
        for cell in row:
            temp_list.append(cell.value)
        if sign==1:
            head=temp_list
        else:
            data_dic[sign] = temp_list
        sign += 1
    import pandas as pd
    df= pd.DataFrame.from_dict(data_dic, orient='index', columns=head)
    return df


def read_excel(excel_name):
  workbook=xlrd.open_workbook(excel_name)
  # 获取所有sheet
  print(workbook.sheet_names()) # [u'sheet1', u'ws']
  ws =workbook.sheet_by_index(0)  #workbook.sheet_names()[1]     workbook.sheet_by_name('Sheet1')
 
  # sheet的名称，行数，列数
  print(ws.name,ws.nrows,ws.ncols)
  # 获取整行和整列的值（数组）
  rows = ws.row_values(3) # 获取第四行内容
  cols = ws.col_values(2) # 获取第三列内容
  # print(rows)
  # print(cols)
  # 获取单元格内容
  # value=ws.cell(1,0).value  #ws.cell_value(1,0)     ws.row(1)[0].value
  #
  # # 获取单元格内容的数据类型
  # print(ws.cell(1,0).ctype)

  # 建立存储数据的字典
  data_dic = {}
  head = []
  # 把数据存到字典中
  for i in range(ws.nrows):
      temp_list = []
      # print("row:", row)
      for j in range(ws.ncols):
          temp_list.append(ws.cell(i,j).value)
      if i == 0:
          head = temp_list
      else:
          data_dic[i] = temp_list
  import pandas as pd
  df = pd.DataFrame.from_dict(data_dic, orient='index', columns=head)
  return df

